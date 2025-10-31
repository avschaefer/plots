import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import re
import numpy as np
from io import BytesIO
from typing import Dict, List, Tuple
import openpyxl
from openpyxl import Workbook

# Page configuration
st.set_page_config(
    page_title="Force Tester Data Visualization",
    layout="wide"
)

def normalize_config_name(name: str) -> str:
    """
    Normalize configuration names by removing common suffix patterns.
    Handles patterns like: syringe-30G-water, syringe-30G-water-1, syringe-30G-water-2
    """
    # Remove trailing digits and hyphens (e.g., -1, -2, -01, etc.)
    pattern = r'-\d+$'
    normalized = re.sub(pattern, '', name)
    return normalized.strip()

def rgb_to_hex(rgb) -> str:
    """
    Convert RGB tuple or string to hex color code.
    Handles both (r, g, b) tuples and 'rgb(r, g, b)' strings.
    """
    if isinstance(rgb, str):
        # Handle 'rgb(r, g, b)' format
        if rgb.startswith('rgb('):
            rgb = rgb[4:-1]  # Remove 'rgb(' and ')'
            r, g, b = map(int, rgb.split(','))
        else:
            # Already a hex string
            return rgb
    elif isinstance(rgb, tuple):
        r, g, b = rgb
    else:
        return rgb
    
    return f"#{r:02x}{g:02x}{b:02x}"

# Color palettes
VIBRANT_COLOR_FIESTA_PALETTE = [
    "#ffbe0b",  # Yellow
    "#fb5607",  # Orange-red
    "#ff006e",  # Hot pink
    "#8338ec",  # Purple
    "#3a86ff"   # Blue
]

VIBRANT_COLOR_BLAST_PALETTE = [
    "#006ba6",  # Deep blue
    "#0496ff",  # Bright blue
    "#ffbc42",  # Golden yellow
    "#d81159",  # Hot pink
    "#8f2d56"   # Deep pink
]

BOLD_HUES_PALETTE = [
    "#f72585",  # Hot pink
    "#7209b7",  # Purple
    "#3a0ca3",  # Deep blue
    "#4361ee",  # Bright blue
    "#4cc9f0"   # Cyan
]

VIBRANT_SUMMER_PALETTE = [
    "#ff595e",  # Coral red
    "#ffca3a",  # Yellow
    "#8ac926",  # Green
    "#1982c4",  # Blue
    "#6a4c93"   # Purple
]

ELECTRIC_DREAMS_PALETTE = [
    "#0015ff",  # Bright blue
    "#ff00a1",  # Hot pink
    "#90fe00",  # Lime green
    "#8400ff",  # Purple
    "#00fff7",  # Cyan
    "#ff7300"   # Orange
]

SUNSET_OCEAN_ORCHID_PALETTE = [
    "#ff595e",  # Coral red
    "#ff924c",  # Orange
    "#ffca3a",  # Yellow
    "#8ac926",  # Green
    "#1982c4",  # Blue
    "#6a4c93"   # Purple
]

def get_palette(palette_name: str):
    """Get color palette by name."""
    palettes = {
        "Vibrant Color Fiesta": VIBRANT_COLOR_FIESTA_PALETTE,
        "Vibrant Color Blast": VIBRANT_COLOR_BLAST_PALETTE,
        "Bold Hues": BOLD_HUES_PALETTE,
        "Vibrant Summer": VIBRANT_SUMMER_PALETTE,
        "Electric Dreams": ELECTRIC_DREAMS_PALETTE,
        "Sunset Ocean Orchid": SUNSET_OCEAN_ORCHID_PALETTE
    }
    return palettes.get(palette_name, VIBRANT_COLOR_FIESTA_PALETTE)

def identify_column_pairs(df_values: pd.DataFrame) -> List[Tuple[int, int, str]]:
    """
    Identify pairs of columns that represent x (travel) and y (force) data.
    Expects DataFrame from "Values Series" sheet with:
    - Row 0: Sample name
    - Row 1: Measurement type ("Standard travel" or "Standard force")
    - Row 2: Units ("mm" or "N")
    - Row 3+: Data values
    
    Returns list of tuples: (x_col_index, y_col_index, sample_name)
    """
    pairs = []
    
    # Check if we have the expected header structure
    if len(df_values) < 3:
        return pairs
    
    num_cols = len(df_values.columns)
    i = 0
    
    while i < num_cols - 1:
        # Get sample name from row 0
        sample_name_x = str(df_values.iloc[0, i])
        sample_name_y = str(df_values.iloc[0, i + 1])
        
        # Get measurement type from row 1
        type_x = str(df_values.iloc[1, i]).strip()
        type_y = str(df_values.iloc[1, i + 1]).strip()
        
        # Check if this is a valid travel/force pair
        if (type_x == "Standard travel" and type_y == "Standard force" and 
            sample_name_x == sample_name_y):
            pairs.append((i, i + 1, sample_name_x))
            i += 2
        else:
            # Skip single column if it doesn't form a pair
            i += 1
    
    return pairs

def get_unique_samples(df_results: pd.DataFrame) -> List[str]:
    """
    Extract unique sample names from "Results Series" sheet.
    Sample names start from row 2, column 0.
    """
    unique_samples = []
    if len(df_results) < 3:
        return unique_samples
    
    # Get sample names from column 0, starting from row 2
    sample_column = df_results.iloc[2:, 0].dropna()
    unique_samples = sample_column.unique().tolist()
    return [str(s).strip() for s in unique_samples if str(s).strip()]


def parse_test_data(df_values: pd.DataFrame, column_pairs: List[Tuple[int, int, str]] = None, 
                    unique_samples: List[str] = None) -> Dict[str, Dict[str, Dict[str, pd.Series]]]:
    """
    Parse the Values Series dataframe and group test samples by configuration.
    Returns a dictionary: {config_name: {sample_name: {'x': x_data, 'y': y_data}}}
    
    Args:
        df_values: DataFrame from "Values Series" sheet (with header rows 0-2)
        column_pairs: Optional list of (x_col_index, y_col_index, sample_name) tuples.
        unique_samples: Optional list of unique sample names from Results Series sheet.
    """
    test_data = {}
    
    # Identify column pairs if not provided
    if column_pairs is None:
        column_pairs = identify_column_pairs(df_values)
    
    if not column_pairs:
        st.error("Could not identify column pairs. Please check the file format.")
        return test_data
    
    # Skip header rows (0-2) and extract data starting from row 3
    data_start_row = 3
    if len(df_values) <= data_start_row:
        st.error("No data rows found. Please check the file format.")
        return test_data
    
    # Extract data (skip header rows)
    df_data = df_values.iloc[data_start_row:].reset_index(drop=True)
    
    for x_col_idx, y_col_idx, sample_name in column_pairs:
        # Use sample name from column pair
        original_name = sample_name
        
        # Normalize configuration name (remove trailing numbers)
        config_name = normalize_config_name(original_name)
        
        # If unique_samples is provided, try to match with authoritative sample name
        if unique_samples:
            # First try exact match
            if original_name in unique_samples:
                base_name = original_name
            else:
                # Try to find matching sample by normalized name
                matching_sample = None
                for us in unique_samples:
                    if normalize_config_name(us) == config_name:
                        matching_sample = us
                        break
                if matching_sample:
                    base_name = matching_sample
                else:
                    base_name = original_name
        else:
            base_name = original_name
        
        if config_name not in test_data:
            test_data[config_name] = {}
        
        # Extract x and y data, removing NaN values
        x_data = df_data.iloc[:, x_col_idx].dropna()
        y_data = df_data.iloc[:, y_col_idx].dropna()
        
        # Align lengths
        min_len = min(len(x_data), len(y_data))
        x_data = x_data.iloc[:min_len]
        y_data = y_data.iloc[:min_len]
        
        if len(x_data) > 0 and len(y_data) > 0:
            # Count existing replicates for this configuration
            existing_count = len(test_data[config_name])
            
            # Create identifier - use base name if first replicate, otherwise add replicate number
            if existing_count == 0:
                replicate_id = base_name
            else:
                replicate_id = f"{base_name}_rep{existing_count + 1}"
            
            test_data[config_name][replicate_id] = {
                'x': x_data,
                'y': y_data
            }
    
    return test_data

def create_filtered_excel(
    df_results: pd.DataFrame,
    df_values: pd.DataFrame,
    test_data: Dict[str, Dict[str, Dict[str, pd.Series]]],
    visibility: Dict[str, bool],
    detected_pairs: List[Tuple[int, int, str]]
) -> BytesIO:
    """
    Create an Excel file with only the selected (visible) groups.
    Maintains the exact format of the original file so it can be re-uploaded.
    
    Args:
        df_results: Original Results Series DataFrame
        df_values: Original Values Series DataFrame
        test_data: Parsed test data dictionary
        visibility: Dictionary of visibility status for each configuration
        detected_pairs: List of (x_col_index, y_col_index, sample_name) tuples
    
    Returns:
        BytesIO object containing the Excel file
    """
    # Get visible configurations
    visible_configs = [config for config in visibility.keys() if visibility.get(config, True)]
    
    if not visible_configs:
        # If no visible configs, return empty file
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Results Series")
        wb.create_sheet("Values Series")
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # Create mapping from original sample names to column pairs
    sample_to_pairs = {}
    for x_col_idx, y_col_idx, sample_name in detected_pairs:
        # Use normalized name as key
        normalized_name = normalize_config_name(sample_name)
        if normalized_name not in sample_to_pairs:
            sample_to_pairs[normalized_name] = []
        sample_to_pairs[normalized_name].append((x_col_idx, y_col_idx, sample_name))
    
    # Get columns to include (all columns for visible configurations)
    columns_to_include = []
    for config_name in visible_configs:
        if config_name in sample_to_pairs:
            # Get all column pairs for this configuration
            for x_col_idx, y_col_idx, orig_sample_name in sample_to_pairs[config_name]:
                columns_to_include.extend([x_col_idx, y_col_idx])
    
    # Remove duplicates while preserving order
    columns_to_include = list(dict.fromkeys(columns_to_include))
    columns_to_include.sort()
    
    # Create Results Series sheet with only visible samples
    results_rows = []
    # Copy first 2 rows (header rows)
    if len(df_results) >= 2:
        for i in range(2):
            row_data = [df_results.iloc[i, j] if j < len(df_results.columns) else None for j in range(max(1, len(df_results.columns)))]
            results_rows.append(row_data)
    
    # Add visible sample names starting from row 2
    added_samples = set()
    for config_name in visible_configs:
        if config_name in sample_to_pairs:
            # Get original sample names for this configuration
            for x_col_idx, y_col_idx, orig_sample_name in sample_to_pairs[config_name]:
                if orig_sample_name not in added_samples:
                    # Find this sample in original Results Series
                    if len(df_results) > 2:
                        for idx in range(2, len(df_results)):
                            orig_name = str(df_results.iloc[idx, 0]).strip() if pd.notna(df_results.iloc[idx, 0]) else ""
                            if orig_name == orig_sample_name:
                                results_rows.append([df_results.iloc[idx, 0]])
                                added_samples.add(orig_sample_name)
                                break
    
    # Create Values Series sheet with only visible columns
    values_rows = []
    
    # Build header rows (0-2)
    max_rows = max(len(df_values), 3)
    header_rows = []
    for row_idx in range(3):
        row_data = []
        for col_idx in columns_to_include:
            if col_idx < len(df_values.columns):
                row_data.append(df_values.iloc[row_idx, col_idx])
            else:
                row_data.append(None)
        header_rows.append(row_data)
    
    # Build data rows (3+)
    data_rows = []
    if len(df_values) > 3:
        max_data_rows = len(df_values) - 3
        for row_idx in range(max_data_rows):
            row_data = []
            for col_idx in columns_to_include:
                if col_idx < len(df_values.columns):
                    value = df_values.iloc[row_idx + 3, col_idx]
                    row_data.append(value)
                else:
                    row_data.append(None)
            data_rows.append(row_data)
    
    # Combine header and data rows
    values_rows = header_rows + data_rows
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create Results Series sheet
    ws_results = wb.create_sheet("Results Series")
    for row_idx, row_data in enumerate(results_rows, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws_results.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Values Series sheet
    ws_values = wb.create_sheet("Values Series")
    for row_idx, row_data in enumerate(values_rows, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws_values.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_chart(
    test_data: Dict[str, Dict[str, Dict[str, pd.Series]]],
    line_types: Dict[str, str],
    line_colors: Dict[str, str],
    sample_names: Dict[str, str],
    visibility: Dict[str, bool],
    title: str,
    subtitle: str,
    x_axis_title: str,
    y_axis_title: str,
    x_min: float,
    x_max: float,
    y_min: float,
    y_max: float,
    color_palette: str = "Vibrant Color Fiesta",
    x_increment: float = None,
    y_increment: float = None,
    font_size: int = 12
):
    """Create a Plotly chart with customizable options."""
    
    fig = go.Figure()
    
    # Define line dash patterns
    dash_patterns = {
        'solid': None,
        'dash': 'dash',
        'dashdot': 'dashdot'
    }
    
    # Generate colors for each configuration
    configs = list(test_data.keys())
    default_colors = get_palette(color_palette)
    
    for config_idx, config_name in enumerate(configs):
        # Skip hidden configurations
        if not visibility.get(config_name, True):
            continue
        
        config_color = line_colors.get(config_name, default_colors[config_idx % len(default_colors)])
        # Convert color to hex if needed for Plotly
        if isinstance(config_color, tuple):
            config_color = rgb_to_hex(config_color)
        
        config_line_type = line_types.get(config_name, 'solid')
        dash_pattern = dash_patterns.get(config_line_type, None)
        
        # Get display name for this configuration
        first_sample = list(test_data[config_name].keys())[0]
        # Check if we have a custom name for any sample in this config
        custom_name = sample_names.get(first_sample, None)
        if custom_name:
            # Remove _repN suffix if present
            if '_rep' in custom_name:
                base_display_name = custom_name.split('_rep')[0]
            else:
                base_display_name = custom_name
        else:
            # Use config_name (which is normalized from sample name)
            base_display_name = config_name
        
        # Add trace for each replicate, but use same name and legendgroup for grouping
        for sample_name, data in test_data[config_name].items():
            fig.add_trace(go.Scatter(
                x=data['x'],
                y=data['y'],
                mode='lines',
                name=base_display_name,  # Same name for all replicates
                legendgroup=config_name,  # Group all replicates together
                showlegend=(sample_name == first_sample),  # Only show in legend once
                line=dict(
                    color=config_color,
                    dash=dash_pattern,
                    width=2
                ),
                hovertemplate=f'<b>{base_display_name}</b><br>' +
                             f'{x_axis_title}: %{{x:.2f}}<br>' +
                             f'{y_axis_title}: %{{y:.2f}}<extra></extra>'
            ))
    
    # Calculate tick values excluding zero
    x_range_min = x_min if x_min is not None else 0
    x_range_max = x_max if x_max is not None else None
    y_range_min = y_min if y_min is not None else 0
    y_range_max = y_max if y_max is not None else None
    
    # Get actual data ranges if limits not set
    if x_range_max is None or y_range_max is None:
        all_x_data = []
        all_y_data = []
        for config_data in test_data.values():
            for sample_data in config_data.values():
                all_x_data.extend(sample_data['x'].tolist())
                all_y_data.extend(sample_data['y'].tolist())
        
        if x_range_max is None:
            x_range_max = max(all_x_data) if all_x_data else 100
        if y_range_max is None:
            y_range_max = max(all_y_data) if all_y_data else 100
    
    # Calculate ticks - exclude zero from labels but include it in the range
    x_increment_val = x_increment if x_increment is not None and x_increment > 0 else 5.0
    y_increment_val = y_increment if y_increment is not None and y_increment > 0 else 5.0
    
    # Generate all ticks including zero
    x_all_ticks = np.arange(x_range_min, x_range_max + x_increment_val, x_increment_val)
    y_all_ticks = np.arange(y_range_min, y_range_max + y_increment_val, y_increment_val)
    
    # Filter out zero for display (but keep the range starting at 0)
    x_ticks = [t for t in x_all_ticks if abs(t) > 1e-10]
    y_ticks = [t for t in y_all_ticks if abs(t) > 1e-10]
    
    # Ensure ranges are explicit (no None values)
    x_range_final = [x_min if x_min is not None else 0, x_max if x_max is not None else x_range_max]
    y_range_final = [y_min if y_min is not None else 0, y_max if y_max is not None else y_range_max]
    
    # Calculate subtitle font size (80% of main font size)
    subtitle_font_size = int(font_size * 0.8)
    
    # Update layout
    fig.update_layout(
        title=dict(
            text=f'<b>{title}</b><br><span style="font-size:{subtitle_font_size}px">{subtitle}</span>',
            x=0.5,
            xanchor='center',
            font=dict(color='black', size=font_size)
        ),
        xaxis=dict(
            title=dict(text=x_axis_title, font=dict(color='black', size=font_size)),
            range=x_range_final,
            tickfont=dict(color='black', size=font_size),
            tickcolor='black',
            gridcolor='lightgrey',
            showgrid=False,  # No vertical grid lines
            zeroline=True,
            zerolinecolor='black',
            zerolinewidth=1.5,
            showline=False,  # Don't show axis border line
            tickmode='array',
            tickvals=x_ticks if x_ticks and len(x_ticks) > 0 else None,
            tickformat='.1f'
        ),
        yaxis=dict(
            title=dict(text=y_axis_title, font=dict(color='black', size=font_size)),
            range=y_range_final,
            tickfont=dict(color='black', size=font_size),
            tickcolor='black',
            gridcolor='lightgrey',
            showgrid=True,  # Show horizontal grid lines
            zeroline=True,
            zerolinecolor='black',
            zerolinewidth=1.5,
            showline=False,  # Don't show axis border line
            tickmode='array',
            tickvals=y_ticks if y_ticks and len(y_ticks) > 0 else None,
            tickformat='.1f'
        ),
        margin=dict(l=60, r=60, t=80, b=60),
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=1.01,
            font=dict(color='black', size=font_size)
        ),
        plot_bgcolor='white',
        paper_bgcolor='white',
        hovermode='closest',
        template='simple_white',
        width=None,
        height=600
    )
    
    # Force update axes separately to ensure proper initialization
    fig.update_xaxes(
        range=x_range_final,
        zeroline=True,
        zerolinecolor='black',
        zerolinewidth=1.5
    )
    fig.update_yaxes(
        range=y_range_final,
        zeroline=True,
        zerolinecolor='black',
        zerolinewidth=1.5
    )
    
    return fig

def main():
    st.title("Force Tester Data Visualization")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls']
    )
    
    if uploaded_file is not None:
        try:
            # Load Excel file and check for required sheets
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names
            
            # Check for required sheets
            if "Results Series" not in sheet_names:
                st.error("Required sheet 'Results Series' not found in the Excel file.")
                return
            
            if "Values Series" not in sheet_names:
                st.error("Required sheet 'Values Series' not found in the Excel file.")
                return
            
            # Load both sheets
            df_results = pd.read_excel(xls, sheet_name="Results Series", header=None)
            df_values = pd.read_excel(xls, sheet_name="Values Series", header=None)
            
            if df_values.empty:
                st.error("The 'Values Series' sheet is empty.")
                return
            
            # Get unique samples from Results Series sheet
            unique_samples = get_unique_samples(df_results)
            
            # Detect column pairs from Values Series sheet
            detected_pairs = identify_column_pairs(df_values)
            
            # Parse test data
            test_data = parse_test_data(df_values, detected_pairs, unique_samples)
            
            if not test_data:
                st.error("No test data found. Please check the file format.")
                return
            
            # Get all unique configurations and samples
            all_configs = list(test_data.keys())
            all_samples = []
            for config_name, samples in test_data.items():
                all_samples.extend(samples.keys())
            
            # Initialize session state for customization if not exists
            if 'line_types' not in st.session_state:
                st.session_state.line_types = {}
            if 'line_colors' not in st.session_state:
                st.session_state.line_colors = {}
            if 'sample_names' not in st.session_state:
                st.session_state.sample_names = {}
            if 'visibility' not in st.session_state:
                st.session_state.visibility = {}
            
            # Update session state with defaults for any new configurations
            for config in all_configs:
                if config not in st.session_state.line_types:
                    st.session_state.line_types[config] = 'solid'
                if config not in st.session_state.visibility:
                    st.session_state.visibility[config] = True
            
            # Initialize color palette if not exists
            if 'color_palette' not in st.session_state:
                st.session_state.color_palette = 'Vibrant Color Fiesta'
            if 'previous_palette' not in st.session_state:
                st.session_state.previous_palette = st.session_state.color_palette
            
            # Configure Test Groups Section
            with st.expander("Configure Test Groups", expanded=False):
                # Create header row
                header_col1, header_col2, header_col3, header_col4, header_col5 = st.columns([0.8, 2, 1.2, 1.2, 2.5])
                header_col1.write("**Show**")
                header_col2.write("**Config**")
                header_col3.write("**Type**")
                header_col4.write("**Color**")
                header_col5.write("**Legend**")
                
                for config_name in all_configs:
                    # Create columns for each configuration
                    col1, col2, col3, col4, col5 = st.columns([0.8, 2, 1.2, 1.2, 2.5])
                    
                    # Column 1: Visibility checkbox
                    is_visible = col1.checkbox(
                        "",
                        value=st.session_state.visibility.get(config_name, True),
                        key=f"visibility_{config_name}",
                        label_visibility="collapsed"
                    )
                    st.session_state.visibility[config_name] = is_visible
                    
                    # Column 2: Configuration name
                    col2.write(config_name)
                    
                    # Column 3: Line type
                    line_type = col3.selectbox(
                        "",
                        options=['solid', 'dash', 'dashdot'],
                        index=['solid', 'dash', 'dashdot'].index(
                            st.session_state.line_types.get(config_name, 'solid')
                        ),
                        key=f"line_type_{config_name}",
                        label_visibility="collapsed"
                    )
                    st.session_state.line_types[config_name] = line_type
                    
                    # Column 4: Line color
                    # Get default color from selected palette
                    selected_palette = get_palette(st.session_state.color_palette)
                    palette_idx = all_configs.index(config_name) % len(selected_palette)
                    palette_default_color = selected_palette[palette_idx]
                    
                    # Check if config has a manually set color
                    stored_color = st.session_state.line_colors.get(config_name, None)
                    
                    # Use palette default if no color is set
                    if stored_color is None:
                        default_color = palette_default_color
                    else:
                        # Convert stored color to hex
                        default_color = rgb_to_hex(stored_color)
                    
                    color = col4.color_picker(
                        "",
                        value=default_color,
                        key=f"color_{config_name}",
                        label_visibility="collapsed"
                    )
                    st.session_state.line_colors[config_name] = color
                    
                    # Column 5: Legend name
                    first_sample = list(test_data[config_name].keys())[0]
                    base_sample_name = first_sample.split('_rep')[0] if '_rep' in first_sample else first_sample
                    current_display_name = st.session_state.sample_names.get(first_sample, base_sample_name)
                    # Remove _rep suffix if present in stored name
                    if '_rep' in current_display_name:
                        current_display_name = current_display_name.split('_rep')[0]
                    
                    new_name = col5.text_input(
                        "",
                        value=current_display_name,
                        key=f"name_{config_name}",
                        label_visibility="collapsed"
                    )
                    # Apply same name to all replicates in this configuration
                    for sample_name in test_data[config_name].keys():
                        st.session_state.sample_names[sample_name] = new_name
            
            # Chart Configuration Section
            st.subheader("Chart Configuration")
            
            # Chart Labels Section
            with st.expander("Chart Labels", expanded=False):
                chart_title = st.text_input("Title", value="Force vs Travel", key="chart_title")
                chart_subtitle = st.text_input("Subtitle", value="", key="chart_subtitle")
                col1, col2 = st.columns(2)
                x_axis_title = col1.text_input("X Axis Title", value="Travel (mm)", key="x_axis_title")
                y_axis_title = col2.text_input("Y Axis Title", value="Force (N)", key="y_axis_title")
                
                # Font size selector
                if 'font_size' not in st.session_state:
                    st.session_state.font_size = 12
                font_size = st.number_input("Font Size", min_value=8, max_value=24, value=st.session_state.font_size, step=1, help="Font size for all chart text (title, labels, legend)")
                st.session_state.font_size = font_size
            
            # Scale Settings Section
            with st.expander("Scale Settings", expanded=False):
                use_custom_x_scale = st.checkbox("Custom X Scale")
                x_min = None
                x_max = None
                x_increment = None
                if use_custom_x_scale:
                    col1, col2 = st.columns(2)
                    x_min = col1.number_input("X Min", value=None, format="%.2f")
                    x_max = col2.number_input("X Max", value=None, format="%.2f")
                    x_increment = st.number_input("X Axis Increment", min_value=0.1, value=5.0, step=0.1, format="%.1f", help="Spacing between ticks on x-axis")
                
                use_custom_y_scale = st.checkbox("Custom Y Scale")
                y_min = None
                y_max = None
                y_increment = None
                if use_custom_y_scale:
                    col1, col2 = st.columns(2)
                    y_min = col1.number_input("Y Min", value=None, format="%.2f")
                    y_max = col2.number_input("Y Max", value=None, format="%.2f")
                    y_increment = st.number_input("Y Axis Increment", min_value=0.1, value=5.0, step=0.1, format="%.1f", help="Spacing between ticks on y-axis")
            
            st.markdown("---")
            
            # Color palette selector above chart
            available_palettes = [
                "Vibrant Color Fiesta",
                "Vibrant Color Blast",
                "Bold Hues",
                "Vibrant Summer",
                "Electric Dreams",
                "Sunset Ocean Orchid"
            ]
            color_palette = st.selectbox(
                "Color Palette",
                options=available_palettes,
                index=available_palettes.index(st.session_state.color_palette) if st.session_state.color_palette in available_palettes else 0,
                key="palette_selector"
            )
            
            # Check if palette changed and reset colors
            if color_palette != st.session_state.color_palette:
                # Palette changed - clear all colors so they update to new palette
                st.session_state.line_colors = {}
                st.session_state.previous_palette = st.session_state.color_palette
                st.session_state.color_palette = color_palette
            else:
                st.session_state.previous_palette = st.session_state.color_palette
                st.session_state.color_palette = color_palette
            
            # Chart below configuration panel
            fig = create_chart(
                test_data,
                st.session_state.line_types,
                st.session_state.line_colors,
                st.session_state.sample_names,
                st.session_state.visibility,
                chart_title,
                chart_subtitle,
                x_axis_title,
                y_axis_title,
                x_min,
                x_max,
                y_min,
                y_max,
                st.session_state.color_palette,
                x_increment,
                y_increment,
                st.session_state.font_size
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Data preview
            with st.expander("Data Preview"):
                st.subheader("Results Series (Sample Names)")
                st.dataframe(df_results.head(20))
                
                st.subheader("Values Series (Raw Data)")
                st.write("**Header rows (0-2):** Sample names, measurement types, and units")
                st.dataframe(df_values.head(20))
                
                if len(df_values) > 20:
                    st.caption(f"Showing first 20 rows of {len(df_values)} total rows")
            
            # Download chart as HTML
            html_str = fig.to_html(include_plotlyjs='cdn')
            st.download_button(
                label="Download Chart as HTML",
                data=html_str,
                file_name="force_tester_chart.html",
                mime="text/html"
            )
            
            # Download Excel data for selected groups
            try:
                excel_data = create_filtered_excel(
                    df_results,
                    df_values,
                    test_data,
                    st.session_state.visibility,
                    detected_pairs
                )
                st.download_button(
                    label="Download Excel Data for Selected Groups",
                    data=excel_data,
                    file_name="filtered_force_tester_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as excel_error:
                st.warning(f"Could not create Excel download: {str(excel_error)}")
            
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            st.exception(e)
    else:
        # Show expected data format
        st.markdown("---")
        st.subheader("Expected Data Format")
        st.markdown("""
        Your Excel file should contain two sheets:
        
        **Sheet 1: "Results Series"**
        - Contains unique sample names starting from row 2, column 0
        
        **Sheet 2: "Values Series"**
        - Row 0: Sample name (repeated for each test)
        - Row 1: Measurement type ("Standard travel" or "Standard force")
        - Row 2: Units ("mm" or "N")
        - Row 3+: Actual data values
        - Columns come in pairs: travel (x) then force (y) for the same sample
        
        Samples with similar names will be automatically grouped together for consistent styling.
        """)

if __name__ == "__main__":
    main()

